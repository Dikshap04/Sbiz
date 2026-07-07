import io
import os
import uuid
from datetime import datetime

from dotenv import load_dotenv
from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, session, url_for

load_dotenv()

_key_preview = os.environ.get("ANTHROPIC_API_KEY")
if _key_preview:
    print(f"[startup] ANTHROPIC_API_KEY found - starts with {_key_preview[:9]!r}, length {len(_key_preview)}")
else:
    print("[startup] ANTHROPIC_API_KEY NOT found in environment.")

from utils import engagement_agent, hr_agent
from utils.resume_parser import allowed_file, extract_resume_text

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key-change-me")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB

SESSIONS: dict[str, dict] = {}
MAX_CANDIDATES = 10


def get_session_state() -> dict:
    sid = session.get("hr_sid")
    if not sid:
        sid = str(uuid.uuid4())
        session["hr_sid"] = sid
    return SESSIONS.setdefault(sid, {
        "position": {"number": "", "title": "", "location": ""},
        "pqs": [""] * 6,
        "candidates": [],      # unified pool — each has text, screening results, chat history
        "group_history": [],   # group chat history
    })


# ── Pages ────────────────────────────────────────────────────────────

@app.route("/healthz")
def healthz():
    return {"status": "ok"}, 200

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/submit", methods=["POST"])
def submit():
    name = (request.form.get("name") or "").strip()
    flash(f"Thanks for reaching out, {name} — we'll be in touch soon." if name else "Please enter your name.")
    return redirect(url_for("home"))

@app.route("/hr")
def hr():
    state = get_session_state()
    return render_template("hr.html",
        position=state["position"],
        pqs=state["pqs"],
        candidates=state["candidates"])

# Keep old URLs working
@app.route("/hr-assistant")
def hr_assistant():
    return redirect(url_for("hr"))

@app.route("/screening")
def screening():
    return redirect(url_for("hr"))


# ── Setup ────────────────────────────────────────────────────────────

@app.route("/api/hr/setup", methods=["POST"])
def hr_setup():
    state = get_session_state()
    data = request.get_json(silent=True) or {}
    state["position"] = {
        "number":   (data.get("number")   or "").strip(),
        "title":    (data.get("title")    or "").strip(),
        "location": (data.get("location") or "").strip(),
    }
    state["pqs"] = [(data.get(f"pq{i}") or "").strip() for i in range(1, 7)]
    return jsonify(ok=True)


# ── Upload & screen ──────────────────────────────────────────────────

@app.route("/api/hr/upload", methods=["POST"])
def hr_upload():
    """Upload one resume, parse it, and screen against PQs if any are defined."""
    state = get_session_state()

    if len(state["candidates"]) >= MAX_CANDIDATES:
        return jsonify(ok=False, error=f"Maximum {MAX_CANDIDATES} resumes per session."), 400

    file = request.files.get("resume")
    if not file or file.filename == "":
        return jsonify(ok=False, error="No file selected."), 400
    if not allowed_file(file.filename):
        return jsonify(ok=False, error="Please upload a PDF, DOCX, or TXT file."), 400

    try:
        text, truncated = extract_resume_text(file.filename, file.read())
    except ValueError as exc:
        return jsonify(ok=False, error=str(exc)), 400
    except Exception:
        return jsonify(ok=False, error="Couldn't read that file."), 400

    active_pqs = [pq for pq in state["pqs"] if pq]

    # Screen against PQs if defined, otherwise just store the resume
    if active_pqs:
        try:
            screening = hr_agent.screen_candidate(text, file.filename, active_pqs)
        except RuntimeError as exc:
            return jsonify(ok=False, error=str(exc)), 500
        except Exception:
            return jsonify(ok=False, error="The AI service had a problem. Please try again."), 502
    else:
        screening = {"name": file.filename, "min_quals": "—", "pqs": []}

    candidate = {
        "id": str(uuid.uuid4()),
        "filename": file.filename,
        "text": text,
        "truncated": truncated,
        "word_count": len(text.split()),
        "name": screening.get("name", file.filename),
        "min_quals": screening.get("min_quals", "—"),
        "pqs": screening.get("pqs", []),
        "chat_history": [],
        # Engagement agent fields
        "email": engagement_agent.extract_email(text),
        "engaged": False,
        "joining_date": "",
        "engagement_log": [],
    }
    state["candidates"].append(candidate)
    state["group_history"] = []  # reset group chat when candidates change

    return jsonify(ok=True, candidate={
        k: candidate[k] for k in ["id", "filename", "name", "min_quals", "pqs", "word_count", "truncated"]
    })


@app.route("/api/hr/remove", methods=["POST"])
def hr_remove():
    state = get_session_state()
    data = request.get_json(silent=True) or {}
    cid = data.get("id")
    state["candidates"] = [c for c in state["candidates"] if c["id"] != cid]
    state["group_history"] = []
    return jsonify(ok=True)


# ── Chat ─────────────────────────────────────────────────────────────

@app.route("/api/hr/ask", methods=["POST"])
def hr_ask():
    state = get_session_state()
    data = request.get_json(silent=True) or {}
    message = (data.get("message") or "").strip()
    mode = data.get("mode", "single")      # "single" or "group"
    candidate_id = data.get("candidate_id")

    if not message:
        return jsonify(ok=False, error="Type a question first."), 400
    if not state["candidates"]:
        return jsonify(ok=False, error="Upload at least one resume first."), 400

    try:
        if mode == "group":
            reply = hr_agent.ask_group(state["candidates"], state["group_history"], message)
            state["group_history"].append({"role": "user", "content": message})
            state["group_history"].append({"role": "assistant", "content": reply})
        else:
            # single mode
            cand = next((c for c in state["candidates"] if c["id"] == candidate_id), None)
            if not cand:
                cand = state["candidates"][0]
            reply = hr_agent.ask_single(cand, cand["chat_history"], message)
            cand["chat_history"].append({"role": "user", "content": message})
            cand["chat_history"].append({"role": "assistant", "content": reply})
    except RuntimeError as exc:
        return jsonify(ok=False, error=str(exc)), 500
    except Exception:
        return jsonify(ok=False, error="The AI service had a problem. Please try again."), 502

    return jsonify(ok=True, reply=reply)


# ── Engagement agent ─────────────────────────────────────────────────

@app.route("/engagement")
def engagement():
    state = get_session_state()
    # Slim view for the client — no full resume text needed in the page.
    slim = [{
        "id": c["id"],
        "name": c.get("name", c["filename"]),
        "filename": c["filename"],
        "email": c.get("email", ""),
        "engaged": c.get("engaged", False),
        "joining_date": c.get("joining_date", ""),
        "engagement_log": c.get("engagement_log", []),
    } for c in state["candidates"]]
    return render_template("engagement.html",
        position=state["position"],
        candidates=slim,
        update_types=engagement_agent.UPDATE_TYPES)


@app.route("/api/engagement/update", methods=["POST"])
def engagement_update():
    """Toggle a candidate's engaged status, or save their email / joining date."""
    state = get_session_state()
    data = request.get_json(silent=True) or {}
    cand = next((c for c in state["candidates"] if c["id"] == data.get("id")), None)
    if not cand:
        return jsonify(ok=False, error="Candidate not found."), 404

    if "email" in data:
        cand["email"] = (data.get("email") or "").strip()
    if "joining_date" in data:
        cand["joining_date"] = (data.get("joining_date") or "").strip()
    if "engaged" in data:
        cand["engaged"] = bool(data.get("engaged"))
    return jsonify(ok=True)


@app.route("/api/engagement/draft", methods=["POST"])
def engagement_draft():
    """Ask Claude for a personalised update draft for one candidate."""
    state = get_session_state()
    data = request.get_json(silent=True) or {}
    cand = next((c for c in state["candidates"] if c["id"] == data.get("id")), None)
    if not cand:
        return jsonify(ok=False, error="Candidate not found."), 404

    update_type = data.get("update_type", "custom")
    custom_notes = (data.get("custom_notes") or "").strip()

    try:
        draft = engagement_agent.draft_update(
            candidate=cand,
            position=state["position"],
            update_type=update_type,
            joining_date=cand.get("joining_date", ""),
            custom_notes=custom_notes,
        )
    except RuntimeError as exc:
        return jsonify(ok=False, error=str(exc)), 500
    except Exception:
        return jsonify(ok=False, error="The AI service had a problem. Please try again."), 502

    return jsonify(ok=True, subject=draft["subject"], body=draft["body"])


@app.route("/api/engagement/send", methods=["POST"])
def engagement_send():
    """Send (or simulate) the reviewed update and log it on the candidate's timeline."""
    state = get_session_state()
    data = request.get_json(silent=True) or {}
    cand = next((c for c in state["candidates"] if c["id"] == data.get("id")), None)
    if not cand:
        return jsonify(ok=False, error="Candidate not found."), 404

    email = (cand.get("email") or "").strip()
    if not email:
        return jsonify(ok=False, error="Add an email address for this candidate first."), 400

    subject = (data.get("subject") or "").strip()
    body = (data.get("body") or "").strip()
    if not subject or not body:
        return jsonify(ok=False, error="Generate or write a subject and message first."), 400

    update_type = data.get("update_type") or "custom"
    sent_ok, status = engagement_agent.send_email(email, subject, body)
    if not sent_ok:
        return jsonify(ok=False, error=f"Couldn't send the email: {status}"), 502

    label = next((t["label"] for t in engagement_agent.UPDATE_TYPES if t["id"] == update_type), "Update")
    entry = {
        "type": update_type,
        "label": label,
        "subject": subject,
        "body": body,
        "sent_at": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "simulated": status == "simulated",
        "direction": "outbound",
    }
    cand.setdefault("engagement_log", []).append(entry)

    return jsonify(ok=True, entry=entry, simulated=(status == "simulated"))


@app.route("/api/engagement/check_replies", methods=["POST"])
def engagement_check_replies():
    """Poll the configured inbox for replies from engaged candidates and
    log any new ones onto their timelines."""
    state = get_session_state()
    candidate_emails = {
        c["email"].strip().lower(): c["id"]
        for c in state["candidates"] if c.get("email")
    }
    seen_ids = state.setdefault("seen_reply_ids", set())

    try:
        replies = engagement_agent.check_replies(candidate_emails, seen_ids)
    except RuntimeError as exc:
        return jsonify(ok=False, error=str(exc)), 400
    except Exception as exc:
        return jsonify(ok=False, error=f"Couldn't check the inbox: {exc}"), 502

    new_entries = []
    for r in replies:
        cand = next((c for c in state["candidates"] if c["id"] == r["candidate_id"]), None)
        if not cand:
            continue
        entry = {
            "type": "reply",
            "label": "Candidate reply",
            "subject": r["subject"],
            "body": r["body"],
            "sent_at": r["received_at"],
            "simulated": False,
            "direction": "inbound",
        }
        cand.setdefault("engagement_log", []).append(entry)
        new_entries.append({"candidate_id": cand["id"], "entry": entry})

    return jsonify(ok=True, new_replies=new_entries, checked=len(new_entries))


# ── Export ───────────────────────────────────────────────────────────

@app.route("/api/hr/export")
def hr_export():
    """Export two-sheet Excel: Screening Summary + Chat Log."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    state = get_session_state()
    candidates = state["candidates"]
    pqs = state["pqs"]
    pos = state["position"]

    if not candidates:
        return jsonify(ok=False, error="Upload at least one resume first."), 400

    wb = openpyxl.Workbook()

    # ── colours & helpers ────────────────────────────────────────────
    HEADER_BG = "D9D9D9"
    TITLE_BG  = "BDD7EE"
    YES_C     = "70AD47"
    NO_C      = "FF0000"
    DARK      = "1D2B36"
    ACCENT    = "C97B3D"
    thin      = Side(style="thin", color="000000")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(ws, row, col, value, bold=False, bg=None, align="center", fg="000000", wrap=True, size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, size=size, color=fg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.border = bdr
        return c

    # ════════════════════════════════════════════════════════════════
    # SHEET 1 — Screening Summary
    # ════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Screening Summary"

    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 14
    for l in ["C","D","E","F","G","H"]:
        ws1.column_dimensions[l].width = 12
    ws1.column_dimensions["I"].width = 10
    ws1.column_dimensions["J"].width = 14
    ws1.column_dimensions["K"].width = 26
    TC = 11

    # Title row
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TC)
    c = ws1.cell(row=1, column=1, value="Applicant Screening Summary Sheet")
    c.font = Font(name="Arial", bold=True, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=TITLE_BG)
    c.border = bdr
    ws1.row_dimensions[1].height = 22

    # Position row
    ws1.row_dimensions[2].height = 18
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    hcell(ws1, 2, 1, f"Position #: {pos['number']}", bold=True, bg=HEADER_BG, align="left")
    ws1.merge_cells(start_row=2, start_column=4, end_row=2, end_column=7)
    hcell(ws1, 2, 4, f"Position Title: {pos['title']}", bold=True, bg=HEADER_BG, align="left")
    ws1.merge_cells(start_row=2, start_column=8, end_row=2, end_column=TC)
    hcell(ws1, 2, 8, f"Position Location: {pos['location']}", bold=True, bg=HEADER_BG, align="left")

    # Column headers rows 3-4
    ws1.row_dimensions[3].height = 30
    ws1.row_dimensions[4].height = 30
    headers = [
        (3,1,4,1,"APPLICANT NAME\nFirst, Last"),
        (3,2,4,2,"Met Minimum\nQualifications"),
        (3,3,4,3,"Met Preferred\nQualification #1"),
        (3,4,4,4,"Met Preferred\nQualification #2"),
        (3,5,4,5,"Met Preferred\nQualification #3"),
        (3,6,4,6,"Met Preferred\nQualification #4"),
        (3,7,4,7,"Met Preferred\nQualification #5"),
        (3,8,4,8,"Met Preferred\nQualification #6"),
        (3,9,4,9,"Total # of PQs Met"),
        (3,10,4,10,"Applicant was\ninterviewed"),
        (3,11,4,11,"Comments regarding\napplicant qualification"),
    ]
    for r1,c1,r2,c2,label in headers:
        ws1.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        hcell(ws1, r1, c1, label, bold=True, bg=HEADER_BG)

    # Candidate rows
    for i, cand in enumerate(candidates):
        row = 5 + i
        ws1.row_dimensions[row].height = 18
        pq_results = cand.get("pqs", [])
        total_yes = sum(1 for v in pq_results if v == "Yes")
        hcell(ws1, row, 1, cand.get("name", cand["filename"]), align="left", bg="FFFFFF")
        min_val = cand.get("min_quals", "—")
        c = hcell(ws1, row, 2, min_val, bg="FFFFFF")
        c.font = Font(name="Arial", size=10, color=YES_C if min_val=="Yes" else (NO_C if min_val=="No" else "000000"))
        for pq_idx in range(6):
            col = 3 + pq_idx
            val = pq_results[pq_idx] if pq_idx < len(pq_results) else ""
            c = hcell(ws1, row, col, val, bg="FFFFFF")
            if val in ("Yes","No"):
                c.font = Font(name="Arial", size=10, color=YES_C if val=="Yes" else NO_C)
        hcell(ws1, row, 9, total_yes, bg="FFFFFF")
        hcell(ws1, row, 10, "", bg="FFFFFF")
        hcell(ws1, row, 11, "", bg="FFFFFF")

    # Empty buffer rows
    for extra in range(5):
        row = 5 + len(candidates) + extra
        ws1.row_dimensions[row].height = 18
        for col in range(1, TC+1):
            hcell(ws1, row, col, "", bg="FFFFFF")

    # PQ legend
    leg = 5 + len(candidates) + 7
    ws1.merge_cells(start_row=leg, start_column=1, end_row=leg, end_column=TC)
    c = ws1.cell(row=leg, column=1,
        value="List of Preferred Qualifications and additional criteria used to screen applicants.")
    c.font = Font(name="Arial", size=9, italic=True)
    for i, pq_label in enumerate(pqs):
        lr = leg + 1 + i
        ws1.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=TC)
        c = ws1.cell(row=lr, column=1, value=f"PQ #{i+1}: {pq_label}")
        c.font = Font(name="Arial", bold=bool(pq_label), size=10)
        c.border = bdr

    ws1.freeze_panes = "A5"

    # ════════════════════════════════════════════════════════════════
    # SHEET 2 — Chat Log
    # ════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Chat Log")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 80

    # Header
    for col, label in enumerate(["Candidate", "Timestamp", "Role", "Message"], start=1):
        c = ws2.cell(row=1, column=col, value=label)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr

    now_str = datetime.now().strftime("%d %b %Y %H:%M")
    current_row = 2

    # Per-candidate chat histories
    for cand in candidates:
        cname = cand.get("name", cand["filename"])
        history = cand.get("chat_history", [])
        if not history:
            continue
        for turn in history:
            role = "Question" if turn["role"] == "user" else "Answer"
            bg = "F2F4F2" if turn["role"] == "user" else "FFFFFF"
            for col, val in enumerate([cname, now_str, role, turn["content"]], start=1):
                c = ws2.cell(row=current_row, column=col, value=val)
                c.font = Font(name="Arial", size=10,
                    color=ACCENT if role=="Question" and col==3 else "000000")
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border = bdr
            current_row += 1

    # Group chat history
    group_history = state.get("group_history", [])
    for turn in group_history:
        role = "Question" if turn["role"] == "user" else "Answer"
        bg = "EDF4FB" if turn["role"] == "user" else "FFFFFF"
        for col, val in enumerate(["[Group]", now_str, role, turn["content"]], start=1):
            c = ws2.cell(row=current_row, column=col, value=val)
            c.font = Font(name="Arial", size=10,
                color=ACCENT if role=="Question" and col==3 else "000000")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = bdr
        current_row += 1

    if current_row == 2:
        ws2.cell(row=2, column=1, value="No chat history yet.").font = Font(name="Arial", italic=True, color="888888")

    ws2.freeze_panes = "A2"

    # ════════════════════════════════════════════════════════════════
    # SHEET 3 — Engagement Log
    # ════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Engagement Log")
    widths3 = [22, 24, 14, 22, 30, 18, 12, 60]
    for col, w in zip("ABCDEFGH", widths3):
        ws3.column_dimensions[col].width = w

    for col, label in enumerate(
        ["Candidate", "Email", "Joining Date", "Update Type", "Subject", "Sent At", "Channel", "Message"],
        start=1,
    ):
        c = ws3.cell(row=1, column=col, value=label)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr

    row3 = 2
    for cand in candidates:
        cname = cand.get("name", cand["filename"])
        cemail = cand.get("email", "")
        cjoin = cand.get("joining_date", "")
        for entry in cand.get("engagement_log", []):
            if entry.get("direction") == "inbound":
                channel = "Reply (inbound)"
            elif entry.get("simulated"):
                channel = "Simulated"
            else:
                channel = "Sent"
            values = [
                cname, cemail, cjoin, entry.get("label", entry.get("type", "")),
                entry.get("subject", ""), entry.get("sent_at", ""), channel, entry.get("body", ""),
            ]
            for col, val in enumerate(values, start=1):
                c = ws3.cell(row=row3, column=col, value=val)
                c.font = Font(name="Arial", size=10,
                    color=ACCENT if channel != "Sent" and col == 7 else "000000")
                c.alignment = Alignment(wrap_text=(col == 8), vertical="top")
                c.border = bdr
            row3 += 1

    if row3 == 2:
        ws3.cell(row=2, column=1, value="No engagement updates sent yet.").font = Font(
            name="Arial", italic=True, color="888888")

    ws3.freeze_panes = "A2"

    # ── Save ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"HR_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
