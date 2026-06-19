import io
import json
import os
import uuid
from datetime import datetime

from dotenv import load_dotenv
from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, session, url_for

load_dotenv()

_key_preview = os.environ.get("ANTHROPIC_API_KEY")
if _key_preview:
    print(f"[startup] ANTHROPIC_API_KEY found - starts with {_key_preview[:9]!r}, length {len(_key_preview)}")
else:
    print("[startup] ANTHROPIC_API_KEY NOT found in environment.")

from utils import hr_agent
from utils.resume_parser import allowed_file, extract_resume_text

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key-change-me")
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024

SESSIONS: dict[str, dict] = {}

MAX_CANDIDATES = 20  # screening sheet rows


def get_session_state() -> dict:
    sid = session.get("hr_sid")
    if not sid:
        sid = str(uuid.uuid4())
        session["hr_sid"] = sid
    return SESSIONS.setdefault(sid, {
        "a": None, "b": None, "history": [],
        # screening sheet state
        "position": {"number": "", "title": "", "location": ""},
        "pqs": [""] * 6,           # up to 6 preferred qualifications
        "candidates": [],           # list of screened candidate dicts
    })


@app.route("/healthz")
def healthz():
    return {"status": "ok"}, 200


@app.route("/")
def home():
    return render_template("index.html")


@app.route("/about")
def about():
    return render_template("about.html")


@app.route("/submit", methods=["POST"])
def submit():
    name = (request.form.get("name") or "").strip()
    if name:
        flash(f"Thanks for reaching out, {name} — we'll be in touch soon.")
    else:
        flash("Please enter your name.")
    return redirect(url_for("home"))


@app.route("/hr-assistant")
def hr_assistant():
    state = get_session_state()
    return render_template("hr_assistant.html", resume_a=state["a"], resume_b=state["b"])


@app.route("/screening")
def screening():
    state = get_session_state()
    return render_template(
        "screening.html",
        position=state["position"],
        pqs=state["pqs"],
        candidates=state["candidates"],
    )


# ── Resume upload (shared between chat and screening) ────────────────
@app.route("/api/hr/upload", methods=["POST"])
def hr_upload():
    state = get_session_state()
    slot = request.form.get("slot")
    if slot not in ("a", "b"):
        return jsonify(ok=False, error="Invalid upload slot."), 400
    file = request.files.get("resume")
    if not file or file.filename == "":
        return jsonify(ok=False, error="No file selected."), 400
    if not allowed_file(file.filename):
        return jsonify(ok=False, error="Please upload a PDF, DOCX, or TXT file."), 400
    try:
        text, truncated = extract_resume_text(file.filename, file.read())
    except ValueError as exc:
        return jsonify(ok=False, error=str(exc)), 400
    except Exception:
        return jsonify(ok=False, error="Couldn't process that file. Please try a different one."), 400
    state[slot] = {"filename": file.filename, "text": text}
    state["history"] = []
    return jsonify(ok=True, slot=slot, filename=file.filename,
                   word_count=len(text.split()), truncated=truncated)


@app.route("/api/hr/clear", methods=["POST"])
def hr_clear():
    state = get_session_state()
    data = request.get_json(silent=True) or {}
    slot = data.get("slot")
    if slot in ("a", "b"):
        state[slot] = None
    else:
        state["a"] = None
        state["b"] = None
    state["history"] = []
    return jsonify(ok=True)


@app.route("/api/hr/ask", methods=["POST"])
def hr_ask():
    state = get_session_state()
    if not state["a"]:
        return jsonify(ok=False, error="Upload at least one resume first."), 400
    data = request.get_json(silent=True) or {}
    message = (data.get("message") or "").strip()
    if not message:
        return jsonify(ok=False, error="Type a question first."), 400
    try:
        reply = hr_agent.ask(state["a"], state["b"], state["history"], message)
    except RuntimeError as exc:
        return jsonify(ok=False, error=str(exc)), 500
    except Exception:
        return jsonify(ok=False, error="The AI service had a problem. Please try again."), 502
    state["history"].append({"role": "user", "content": message})
    state["history"].append({"role": "assistant", "content": reply})
    return jsonify(ok=True, reply=reply)


# ── Screening sheet routes ───────────────────────────────────────────

@app.route("/api/screening/setup", methods=["POST"])
def screening_setup():
    """Save position info and PQ definitions."""
    state = get_session_state()
    data = request.get_json(silent=True) or {}
    state["position"] = {
        "number": (data.get("number") or "").strip(),
        "title":  (data.get("title")  or "").strip(),
        "location": (data.get("location") or "").strip(),
    }
    state["pqs"] = [(data.get(f"pq{i}") or "").strip() for i in range(1, 7)]
    return jsonify(ok=True)


@app.route("/api/screening/upload", methods=["POST"])
def screening_upload():
    """Upload a resume for screening, auto-evaluate against PQs with Claude."""
    state = get_session_state()

    if len(state["candidates"]) >= MAX_CANDIDATES:
        return jsonify(ok=False, error=f"Maximum {MAX_CANDIDATES} candidates per session."), 400

    active_pqs = [pq for pq in state["pqs"] if pq]
    if not active_pqs:
        return jsonify(ok=False, error="Define at least one Preferred Qualification before screening."), 400

    file = request.files.get("resume")
    if not file or file.filename == "":
        return jsonify(ok=False, error="No file selected."), 400
    if not allowed_file(file.filename):
        return jsonify(ok=False, error="Please upload a PDF, DOCX, or TXT file."), 400

    try:
        text, _ = extract_resume_text(file.filename, file.read())
    except ValueError as exc:
        return jsonify(ok=False, error=str(exc)), 400
    except Exception:
        return jsonify(ok=False, error="Couldn't read that file."), 400

    # Ask Claude to evaluate this resume against every PQ in one call
    try:
        result = hr_agent.screen_candidate(text, file.filename, active_pqs)
    except RuntimeError as exc:
        return jsonify(ok=False, error=str(exc)), 500
    except Exception:
        return jsonify(ok=False, error="The AI service had a problem. Please try again."), 502

    state["candidates"].append(result)
    return jsonify(ok=True, candidate=result)


@app.route("/api/screening/remove", methods=["POST"])
def screening_remove():
    state = get_session_state()
    data = request.get_json(silent=True) or {}
    idx = data.get("index")
    if isinstance(idx, int) and 0 <= idx < len(state["candidates"]):
        state["candidates"].pop(idx)
    return jsonify(ok=True)


@app.route("/api/screening/export")
def screening_export():
    """Export the Applicant Screening Summary Sheet as .xlsx."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    state = get_session_state()
    candidates = state["candidates"]
    pqs = state["pqs"]
    pos = state["position"]
    active_pq_count = sum(1 for p in pqs if p)

    if not candidates:
        return jsonify(ok=False, error="No candidates screened yet."), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screening Summary"

    # ── helpers ──────────────────────────────────────────────────────
    HEADER_BG  = "D9D9D9"   # light grey — matches the screenshot
    TITLE_BG   = "BDD7EE"   # light blue for the main title row
    YES_COLOR  = "70AD47"   # green text for Yes
    NO_COLOR   = "FF0000"   # red text for No
    BORDER_COLOR = "000000"

    thin = Side(style="thin", color=BORDER_COLOR)
    full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(row, col, value, bold=False, bg=None, wrap=True, align="center", fg="000000", size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, size=size, color=fg)
        c.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.border = full_border
        return c

    # ── Column widths ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 24   # Applicant Name
    ws.column_dimensions["B"].width = 14   # Min Quals
    for col_letter in ["C","D","E","F","G","H"]:
        ws.column_dimensions[col_letter].width = 12  # PQ #1-6
    ws.column_dimensions["I"].width = 10   # Total PQs
    ws.column_dimensions["J"].width = 14   # Interviewed
    ws.column_dimensions["K"].width = 26   # Comments

    total_cols = 11  # A through K

    # ── Row 1: Title ─────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value="Applicant Screening Summary Sheet")
    c.font = Font(name="Arial", bold=True, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=TITLE_BG)
    c.border = full_border

    # ── Row 2: Position info ─────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    hcell(2, 1, f"Position #: {pos['number']}", bold=True, bg=HEADER_BG, align="left")
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=7)
    hcell(2, 4, f"Position Title: {pos['title']}", bold=True, bg=HEADER_BG, align="left")
    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=total_cols)
    hcell(2, 8, f"Position Location: {pos['location']}", bold=True, bg=HEADER_BG, align="left")

    # ── Row 3-4: Column headers (merged for multi-line) ──────────────
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 30

    header_cells = [
        (3, 1, 4, 1, "APPLICANT NAME\nFirst, Last"),
        (3, 2, 4, 2, "Met Minimum\nQualifications"),
        (3, 3, 4, 3, "Met Preferred\nQualification #1"),
        (3, 4, 4, 4, "Met Preferred\nQualification #2"),
        (3, 5, 4, 5, "Met Preferred\nQualification #3"),
        (3, 6, 4, 6, "Met Preferred\nQualification #4"),
        (3, 7, 4, 7, "Met Preferred\nQualification #5"),
        (3, 8, 4, 8, "Met Preferred\nQualification #6"),
        (3, 9, 4, 9, "Total # of PQs Met"),
        (3, 10, 4, 10, "Applicant was\ninterviewed"),
        (3, 11, 4, 11, "Comments regarding\napplicant qualification"),
    ]
    for r1, c1, r2, c2, label in header_cells:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        hcell(r1, c1, label, bold=True, bg=HEADER_BG, wrap=True, align="center")

    # ── Candidate rows (starting row 5) ─────────────────────────────
    for i, cand in enumerate(candidates):
        row = 5 + i
        ws.row_dimensions[row].height = 18
        pq_results = cand.get("pqs", [])
        total_yes = sum(1 for r in pq_results if r == "Yes")

        # Name
        hcell(row, 1, cand.get("name", cand.get("filename", "")),
              align="left", bg="FFFFFF")
        # Met minimum quals (AI determined)
        min_val = cand.get("min_quals", "Yes")
        c = hcell(row, 2, min_val, bg="FFFFFF")
        c.font = Font(name="Arial", size=10,
                      color=YES_COLOR if min_val == "Yes" else NO_COLOR)

        # PQ columns 3-8
        for pq_idx in range(6):
            col = 3 + pq_idx
            if pq_idx < len(pq_results):
                val = pq_results[pq_idx]
                c = hcell(row, col, val, bg="FFFFFF")
                if val in ("Yes", "No"):
                    c.font = Font(name="Arial", size=10,
                                  color=YES_COLOR if val == "Yes" else NO_COLOR)
            else:
                hcell(row, col, "", bg="FFFFFF")

        # Total PQs met
        hcell(row, 9, total_yes, bg="FFFFFF")
        # Interviewed (blank for user to fill)
        hcell(row, 10, "", bg="FFFFFF")
        # Comments (blank for user to fill)
        hcell(row, 11, "", bg="FFFFFF")

    # ── Empty rows for manual additions ────────────────────────────
    for extra in range(5):
        row = 5 + len(candidates) + extra
        ws.row_dimensions[row].height = 18
        for col in range(1, total_cols + 1):
            hcell(row, col, "", bg="FFFFFF")

    # ── PQ legend at the bottom ──────────────────────────────────────
    legend_start = 5 + len(candidates) + 5 + 2
    ws.merge_cells(start_row=legend_start, start_column=1,
                   end_row=legend_start, end_column=total_cols)
    c = ws.cell(row=legend_start, column=1,
                value="List of Preferred Qualifications and additional criteria used to screen applicants (add additional pages as needed).")
    c.font = Font(name="Arial", size=9, italic=True)
    c.alignment = Alignment(wrap_text=True)

    for i, pq_label in enumerate(pqs):
        leg_row = legend_start + 1 + i
        ws.merge_cells(start_row=leg_row, start_column=1,
                       end_row=leg_row, end_column=total_cols)
        display = pq_label if pq_label else ""
        c = ws.cell(row=leg_row, column=1, value=f"PQ #{i+1}: {display}")
        c.font = Font(name="Arial", bold=bool(pq_label), size=10)
        c.border = full_border

    # ── Freeze the header rows ───────────────────────────────────────
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Screening_Summary_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
